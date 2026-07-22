import sqlite3
from pathlib import Path
from openpyxl import load_workbook

EXCEL_PATH = Path("data/MegaUniformes_Base_Datos_Profesional.xlsx")
DB_PATH = Path("inventario.db")


def limpiar(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def main():
    if not EXCEL_PATH.exists():
        print("No se encontró el Excel en backend/data")
        return

    print("Leyendo Excel...")

    wb = load_workbook(EXCEL_PATH, data_only=True)

    print("Hojas disponibles en el Excel:")
    print(wb.sheetnames)

    hoja = wb["Inventario_Web"]

    print("Usando hoja: Inventario_Web")

    headers = [limpiar(cell.value) for cell in hoja[1]]
    print("Columnas encontradas:")
    print(headers)

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    cursor.execute("DROP TABLE IF EXISTS inventario")

    cursor.execute("""
        CREATE TABLE inventario (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            colegio TEXT,
            prenda TEXT,
            genero TEXT,
            talla TEXT,
            stock INTEGER,
            precio REAL,
            descripcion_original TEXT
        )
    """)

    total = 0

    for row in hoja.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        colegio = limpiar(data.get("Colegio"))
        prenda = limpiar(data.get("Prenda"))
        genero = limpiar(data.get("Genero"))
        talla = limpiar(data.get("Talla"))
        descripcion = limpiar(data.get("Descripcion_Original"))

        try:
            stock = int(data.get("Stock") or 0)
        except:
            stock = 0

        try:
            precio = float(data.get("Precio") or 0)
        except:
            precio = 0

        cursor.execute("""
            INSERT INTO inventario 
            (colegio, prenda, genero, talla, stock, precio, descripcion_original)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (colegio, prenda, genero, talla, stock, precio, descripcion))

        total += 1

    conn.commit()
    conn.close()

    print("Base de datos creada correctamente: inventario.db")
    print(f"Total de productos importados: {total}")


if __name__ == "__main__":
    main()